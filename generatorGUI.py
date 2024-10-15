import tkinter as tk
from tkinterdnd2 import TkinterDnD, DND_FILES
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def browse_file(label, side):
    filename = filedialog.askopenfilename(title=f"Select the {side} Excel file", filetypes=[("Excel files", "*.xlsx")])
    if filename:
        label.config(text=filename)
    return filename

def drop_file(event, label):
    filename = event.data
    if filename.endswith('.xlsx'):
        label.config(text=filename)
    else:
        messagebox.showerror("Error", "Please drop an Excel file (.xlsx)")

def clear_file(label):
    label.config(text="Drag and drop or click to upload the file")

def compare_sheets():
    file1 = left_label.cget("text")
    file2 = right_label.cget("text")

    if not file1 or not file2 or "Drag" in file1 or "Drag" in file2:
        messagebox.showerror("Error", "Please select or drop both files before comparing.")
        return

    try:
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)
        comparison = df1.compare(df2, keep_shape=True, keep_equal=True)

        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df1.to_excel(writer, sheet_name='Comparison', index=False)

            wb = load_workbook(output_file)
            ws = wb['Comparison']
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

            for row in range(2, len(df1) + 2):
                for col in range(1, len(df1.columns) + 1):
                    cell_value1 = ws.cell(row=row, column=col).value
                    cell_value2 = df2.iloc[row-2, col-1] if row-2 < len(df2) and col-1 < len(df2.columns) else None
                    if pd.notna(cell_value1) and pd.notna(cell_value2) and cell_value1 != cell_value2:
                        ws.cell(row=row, column=col).fill = red_fill

            wb.save(output_file)

            messagebox.showinfo("Success", f"Comparison complete. Results saved to {output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def animate_opening(app):
    def animate_resize(width):
        current_width = app.winfo_width()
        if current_width < width:
            app.geometry(f"{current_width + 10}x500")
            app.after(10, animate_resize, width)
        else:
            app.geometry(f"{width}x500")
    
    # Start the animation with a desired width
    animate_resize(900)

app = TkinterDnD.Tk()
app.title("Excel Comparator")
app.geometry("200x500")  # Start with a smaller size for animation

# Call the animation function
app.after(100, lambda: animate_opening(app))

app.configure(bg="#f5f5f5")

# Main Frame
main_frame = tk.Frame(app, bg="#f5f5f5")
main_frame.pack(fill=tk.BOTH, expand=True)

# Left Section (File to Compare)
left_frame = tk.Frame(main_frame, padx=20, pady=20, bg="#ffffff", relief=tk.RAISED, borderwidth=2)
left_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

left_label = tk.Label(left_frame, text="Drag and drop or click to upload the file to compare", bg="#f0f0f0", fg="#333333", width=40, height=10, relief=tk.SUNKEN, borderwidth=2, font=("Arial", 12))
left_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
left_label.drop_target_register(DND_FILES)
left_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, left_label))

left_button_frame = tk.Frame(left_frame, bg="#ffffff")
left_button_frame.pack(fill=tk.X, pady=10)

left_button = tk.Button(left_button_frame, text="Upload File", command=lambda: browse_file(left_label, "first"), bg="#007bff", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
left_button.pack(side=tk.LEFT, padx=5)

left_clear_button = tk.Button(left_button_frame, text="Clear", command=lambda: clear_file(left_label), bg="#dc3545", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
left_clear_button.pack(side=tk.LEFT, padx=5)

# Right Section (Comparison File)
right_frame = tk.Frame(main_frame, padx=20, pady=20, bg="#ffffff", relief=tk.RAISED, borderwidth=2)
right_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

right_label = tk.Label(right_frame, text="Drag and drop or click to upload the comparison file", bg="#f0f0f0", fg="#333333", width=40, height=10, relief=tk.SUNKEN, borderwidth=2, font=("Arial", 12))
right_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
right_label.drop_target_register(DND_FILES)
right_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, right_label))

right_button_frame = tk.Frame(right_frame, bg="#ffffff")
right_button_frame.pack(fill=tk.X, pady=10)

right_button = tk.Button(right_button_frame, text="Upload File", command=lambda: browse_file(right_label, "second"), bg="#28a745", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
right_button.pack(side=tk.LEFT, padx=5)

right_clear_button = tk.Button(right_button_frame, text="Clear", command=lambda: clear_file(right_label), bg="#dc3545", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
right_clear_button.pack(side=tk.LEFT, padx=5)

# Compare Button
compare_button = tk.Button(main_frame, text="Compare Excel Sheets", command=compare_sheets, bg="#ff5722", fg="#ffffff", font=("Arial", 12), relief=tk.FLAT)
compare_button.grid(row=1, column=0, columnspan=2, pady=20)

# Configure grid weights
main_frame.grid_rowconfigure(0, weight=1)
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)

app.mainloop()
